import openpyxl
from sklearn.svm import SVC
from sklearn.model_selection import GridSearchCV
from sklearn.feature_selection import RFE
from sklearn import svm
from sklearn import metrics
import numpy as np



#Number of features taken into consideration
Number_of_features=20

#Number of training images taken into consideration
Number_of_training_images=30

#Number of testing images taken into consideration
Number_of_testing_images=12

print("Reading Training data from file")
# Reading X from an excel file X.xlsx
wb = openpyxl.load_workbook('X3.xlsx')
ws=wb.active

X=[[0 for j in range(Number_of_features)] for i in range(Number_of_training_images)]

for i in range(0,Number_of_training_images):
    for j in range(0,Number_of_features):
        X[i][j]=ws[chr(ord('A') + j) + str(i + 1)].value
    print("Reading image : ",i)

print(X)

print("Done reading training data from file")

X1=[1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1]
X2=[2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2]
X3=[3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3]
X4=[4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4]
X5=[5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5,5]
X6=[6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6,6]
X7=[7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7,7]


X_classification =np.append(np.append(np.append(np.append(np.append(np.append(X1,X2),X3),X4),X5),X6),X7)
X_classification=[1,1,1,1,1,1,1,1,1,1,2,2,2,2,2,2,2,2,2,2,3,3,3,3,3,3,3,3,3,3]

print("Reading test data from file")
# Reading Y from an excel file
wb = openpyxl.load_workbook('Y3.xlsx')
ws=wb.active

Y=[[0 for j in range(Number_of_features)] for i in range(Number_of_testing_images)]

for i in range(0,Number_of_testing_images):
    for j in range(0,Number_of_features):
        Y[i][j]=ws[chr(ord('A') + j) + str(i + 1)].value
    print("Reading image : ", i)

print("Done reading test data from file")
Y1=[1,1,1,1,1,1,1,1,1,1]
Y2=[2,2,2,2,2,2,2,2,2,2]
Y3=[3,3,3,3,3,3,3,3,3,3]
Y4=[4,4,4,4,4,4,4,4,4,4]
Y5=[5,5,5,5,5,5,5,5,5,5]
Y6=[6,6,6,6,6,6,6,6,6,6]
Y7=[7,7,7,7,7,7,7,7,7,7]

Y_classification=np.append(np.append(np.append(np.append(np.append(np.append(Y1,Y2),Y3),Y4),Y5),Y6),Y7)

Y_classification=[1,1,1,1,2,2,2,2,3,3,3,3]
"""
# Select number of features to be used
Number_of_features_tobeused=20

#Using linear SVM to assign weights to the features

print("Starting classification")

print("Performing Recursive Feature Elimination on Training")
# Create the RFE object and rank each pixel and using the ranks of the features, select the highest ranked features
svc = SVC(kernel="linear", C=1)
rfe = RFE(estimator=svc, n_features_to_select=Number_of_features_tobeused, step=1)
print("Fitting Training dataset into RFE")
rfe.fit(X,X_classification)
print("Transforming Training dataset ")
X_transformed=rfe.transform(X)

print("Shape of transformed training dataset:",X_transformed.shape)

print("Performing Recursive Feature Elimination on Testing")
svc = SVC(kernel="linear", C=1)
rfe = RFE(estimator=svc, n_features_to_select=Number_of_features_tobeused, step=1)
print("Fitting Testing dataset into RFE")
rfe.fit(Y,Y_classification)
Y_transformed=rfe.transform(Y)

print("Shape of transformed testing dataset:",Y_transformed.shape)

"""

param_grid = {'C': [1e3, 5e3, 1e4, 5e4, 1e5],
              'gamma': [0.0001, 0.0005, 0.001, 0.005, 0.01, 0.1], }
print("Starting classification using SVM")
clf = GridSearchCV(SVC(kernel='poly', class_weight='balanced'), param_grid)

clf = clf.fit(X,X_classification)


Y_test_classification=clf.predict(Y)
accuracyscore=metrics.accuracy_score(Y_classification,Y_test_classification)

print("Accuracy score using SVM: ",accuracyscore)
print("Number of features used:",Number_of_features)

